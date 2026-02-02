"""
Excel 파싱 서비스
발주서 엑셀 파일을 파싱하여 주문 데이터로 변환
"""
from openpyxl import load_workbook
from typing import List, Dict, Optional
import io


# 헤더 인식 키워드
HEADER_KEYWORDS = {
    'recipient': ['수령인', '수취인', '받는분', '받는사람', '수령자', '성명', '이름', '고객명'],
    'phone': ['연락처', '휴대폰', '핸드폰', '전화', '폰번호', '휴대전화', '전화번호'],
    'address': ['주소', '배송지', '배송주소', '도로명', '지번'],
    'product': ['상품', '품명', '품목', '물품', '제품', '물건', '아이템', '내용물'],
    'quantity': ['수량', '개수'],
    'order_date': ['출고요청일', '출고일', '배송일', '요청일'],
    'memo': ['배송메세지', '메모', '요청사항', '비고'],
    'order_no': ['주문번호', '오더번호', '주문'],
    'sender': ['발송인', '보내는분', '송하인'],
    'sender_phone': ['발송인연락처', '보내는분연락처'],
    'sender_address': ['발송인주소', '보내는분주소'],
}


def find_header_row(ws) -> Optional[int]:
    """헤더 행 찾기"""
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), 1):
        if not row:
            continue

        # 헤더 키워드가 2개 이상 있으면 헤더 행으로 판단
        match_count = 0
        for cell in row:
            if cell is None:
                continue
            cell_str = str(cell).strip()
            for keywords in HEADER_KEYWORDS.values():
                if any(kw in cell_str for kw in keywords):
                    match_count += 1
                    break

        if match_count >= 2:
            return row_idx

    return None


def map_columns(header_row: tuple) -> Dict[str, int]:
    """헤더 행에서 컬럼 매핑"""
    mapping = {}

    for col_idx, cell in enumerate(header_row):
        if cell is None:
            continue
        cell_str = str(cell).strip()

        for field, keywords in HEADER_KEYWORDS.items():
            if any(kw in cell_str for kw in keywords):
                if field not in mapping:  # 첫 번째 매칭만 사용
                    mapping[field] = col_idx
                break

    return mapping


def parse_excel(file_content: bytes, vendor_name: str) -> List[Dict]:
    """
    엑셀 파일 파싱

    Args:
        file_content: 엑셀 파일 바이트
        vendor_name: 발주처명

    Returns:
        파싱된 주문 리스트
    """
    wb = load_workbook(io.BytesIO(file_content), data_only=True)
    results = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # 헤더 행 찾기
        header_row_idx = find_header_row(ws)
        if header_row_idx is None:
            continue

        # 헤더에서 컬럼 매핑
        header_row = list(ws.iter_rows(min_row=header_row_idx, max_row=header_row_idx, values_only=True))[0]
        column_mapping = map_columns(header_row)

        if not column_mapping:
            continue

        # 데이터 행 파싱
        for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
            if not any(row):  # 빈 행 스킵
                continue

            order = {
                'vendor_name': vendor_name,
                'recipient': get_cell_value(row, column_mapping.get('recipient')),
                'phone': get_cell_value(row, column_mapping.get('phone')),
                'address': get_cell_value(row, column_mapping.get('address')),
                'product_name': get_cell_value(row, column_mapping.get('product')),
                'quantity': parse_int(get_cell_value(row, column_mapping.get('quantity')), 1),
                'order_date': get_cell_value(row, column_mapping.get('order_date')),
                'memo': get_cell_value(row, column_mapping.get('memo')),
                'order_no': get_cell_value(row, column_mapping.get('order_no')),
                'sender_name': get_cell_value(row, column_mapping.get('sender')),
                'sender_phone': get_cell_value(row, column_mapping.get('sender_phone')),
                'sender_address': get_cell_value(row, column_mapping.get('sender_address')),
            }

            # 최소한 수령인 또는 상품명이 있어야 유효
            if order['recipient'] or order['product_name']:
                results.append(order)

    wb.close()
    return results


def get_cell_value(row: tuple, col_idx: Optional[int]) -> Optional[str]:
    """셀 값 가져오기"""
    if col_idx is None or col_idx >= len(row):
        return None
    value = row[col_idx]
    if value is None:
        return None
    return str(value).strip()


def parse_int(value: Optional[str], default: int = 0) -> int:
    """정수 파싱"""
    if value is None:
        return default
    try:
        # 쉼표 제거 후 파싱
        return int(str(value).replace(',', '').strip())
    except (ValueError, TypeError):
        return default
